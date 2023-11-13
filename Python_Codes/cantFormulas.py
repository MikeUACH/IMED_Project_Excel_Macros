import openpyxl

def contar_referencias_excel(archivo_excel):
    # Abre el archivo Excel
    workbook = openpyxl.load_workbook(archivo_excel)  # Usamos data_only para obtener los valores calculados en lugar de fórmulas

    # Inicializa un diccionario para almacenar la información de las referencias
    referencias = {}

    # Recorre cada hoja en el archivo Excel
    for sheet_name in workbook.sheetnames:
        # Inicializa un conjunto para almacenar las celdas únicas referenciadas en cada hoja
        celdas_referenciadas = set()

        # Obtiene la hoja actual
        hoja_actual = workbook[sheet_name]

        # Recorre todas las celdas en la hoja actual
        for row in hoja_actual.iter_rows():
            for cell in row:
                # Obtiene el contenido de la celda
                contenido_celda = cell.value

                # Verifica si la celda contiene una fórmula (por ejemplo, comienza con '=')
                if isinstance(contenido_celda, str) and contenido_celda.startswith('='):
                    # Divide la fórmula en partes para verificar referencias
                    partes_formula = contenido_celda.split("!")
                    if len(partes_formula) > 1:
                        # Obtiene la referencia de la hoja
                        referencia_hoja = partes_formula[0]
                        # Elimina los caracteres de comillas y el signo igual al principio
                        referencia_hoja = referencia_hoja.strip("'=")

                        # Si la referencia no es la hoja actual, agrega la referencia al conjunto
                        if referencia_hoja != hoja_actual.title:
                            celdas_referenciadas.add(f"En la hoja '{sheet_name}', celda {cell.coordinate} referencia a la hoja '{referencia_hoja}'")

        # Agrega el conjunto de referencias de la hoja al diccionario
        referencias[sheet_name] = celdas_referenciadas

    # Cierra el archivo Excel
    workbook.close()

    # Devuelve el diccionario de referencias
    return referencias

# Ejemplo de uso
archivo_excel = "Q124_BU_Scenario_Flexline_IMED.xlsx"
referencias = contar_referencias_excel(archivo_excel)

# Imprime la información de las referencias
for sheet_name, celdas_referenciadas in referencias.items():
    print(f"En la hoja '{sheet_name}':")
    for referencia in celdas_referenciadas:
        print(referencia)
    print()
